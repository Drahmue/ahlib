import os
import pandas as pd
import msvcrt
from datetime import datetime
import inspect
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import configparser


# --------------------------
# Parquet-Dateien exportieren
# --------------------------
def export_df_to_parquet(df, filename, compression=None, logfile=None, screen=True):
    """
    Exportiert einen DataFrame in eine Parquet-Datei.

    Parameter:
        df (pandas.DataFrame): Der zu exportierende DataFrame.
        filename (str): Der Pfad zur Zieldatei.
        compression (str, optional): Komprimierung ('snappy', 'gzip', 'brotli', None).
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.

    Rückgabe:
        bool: True bei erfolgreichem Export, False bei Fehler.
    """
    try:
        # Typprüfung DataFrame
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Das übergebene Objekt ('df') ist kein gültiger pandas.DataFrame.")
        
        # Prüfung auf leeren DataFrame
        if df.empty:
            raise ValueError("Der DataFrame ist leer und kann nicht exportiert werden.")
        
        # Dateiendung prüfen
        if not filename.endswith('.parquet'):
            raise ValueError(f"Die Datei '{filename}' hat keine '.parquet'-Endung.")
        
        # Verzeichnis erstellen falls nicht vorhanden
        dir_name = os.path.dirname(filename)
        if dir_name and not os.path.exists(dir_name):
            os.makedirs(dir_name, exist_ok=True)
            screen_and_log(f"Info: Verzeichnis '{dir_name}' wurde erstellt.", logfile, screen)
        
        # DataFrame exportieren
        df.to_parquet(filename, compression=compression)
        screen_and_log(f"Info: DataFrame erfolgreich in '{filename}' exportiert.", logfile, screen)
        return True
        
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei oder Verzeichnis nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Eingabe: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Exportieren des DataFrames in '{filename}': {e}", logfile, screen)
        return False

# --------------------------
# Excel-Dateien exportieren
# --------------------------
def export_df_to_excel(df, filename, sheet_name='Sheet1', logfile=None, screen=True):
    """
    Exportiert einen DataFrame in eine Excel-Datei.

    Parameter:
        df (pandas.DataFrame): Der zu exportierende DataFrame.
        filename (str): Der Pfad zur Zieldatei.
        sheet_name (str): Name des Arbeitsblatts (Standard: 'Sheet1').
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.

    Rückgabe:
        bool: True bei erfolgreichem Export, False bei Fehler.
    """
    try:
        # Typprüfung DataFrame
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Das übergebene Objekt ('df') ist kein gültiger pandas.DataFrame.")
        
        # Prüfung auf leeren DataFrame
        if df.empty:
            raise ValueError("Der DataFrame ist leer und kann nicht exportiert werden.")
        
        # Dateiendung prüfen
        if not filename.endswith('.xlsx'):
            raise ValueError(f"Die Datei '{filename}' hat keine '.xlsx'-Endung.")
        
        # Verzeichnis erstellen falls nicht vorhanden
        dir_name = os.path.dirname(filename)
        if dir_name and not os.path.exists(dir_name):
            os.makedirs(dir_name, exist_ok=True)
            screen_and_log(f"Info: Verzeichnis '{dir_name}' wurde erstellt.", logfile, screen)
        
        # DataFrame exportieren - prüfe ob Index als Spalte benötigt wird
        if df.index.name or (hasattr(df.index, 'names') and any(df.index.names)):
            # Index hat einen Namen (z.B. 'date') -> als Spalte exportieren
            df_export = df.reset_index()
            df_export.to_excel(filename, sheet_name=sheet_name, index=False)
        else:
            # Index ist Standard-Index -> ohne Index exportieren
            df.to_excel(filename, sheet_name=sheet_name, index=False)
        screen_and_log(f"Info: DataFrame erfolgreich in '{filename}' exportiert.", logfile, screen)
        return True
        
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei oder Verzeichnis nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Eingabe: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Exportieren des DataFrames in '{filename}': {e}", logfile, screen)
        return False

# -------------------------------
# Exportiert einen 2D Datafram mit Multiindex in eine Pivot-artige EXCEL Tablle 
# -------------------------------
def export_2D_df_to_excel_pivot(df, filename, sheet_name='Sheet1', logfile=None, screen=True):
    """
    Exportiert den übergebenen DataFrame in eine Pivot-Darstellung als Excel-Datei.
    Zeilen enthalten die erste Indexebene, Spalten die zweite Indexebene und die Zellen den Wert.

    Parameter:
        df (DataFrame): Der zu exportierende DataFrame (MultiIndex erwartet).
        filename (str): Der Pfad und Name der Datei, in die der DataFrame exportiert werden soll.
        sheet_name (str): Name des Arbeitsblatts (Standard: 'Sheet1').
        logfile (str, optional): Der Pfad zum Logfile. Wenn None, wird keine Protokollierung ins Logfile durchgeführt.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.

    Rückgabe:
        bool: True bei erfolgreichem Export, False bei Fehler.

    Hinweise:
        - Der DataFrame muss mindestens zwei Indexebenen enthalten.
        - Der Dateiname muss auf ".xlsx" enden.
        - Das Zielverzeichnis muss existieren und beschreibbar sein.
        - Wenn der Export fehlschlägt, wird eine Fehlermeldung protokolliert.
    """
    try:
        # Typprüfung der Eingabeparameter
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Das übergebene Objekt ('df') ist kein gültiger pandas.DataFrame.")
        if not isinstance(filename, str):
            raise ValueError("Der Dateiname ('filename') muss ein String sein.")
        if logfile and not isinstance(logfile, str):
            raise ValueError("Der Logfile-Pfad ('logfile') muss ein String sein.")

        # Überprüfen, ob der Dateiname eine .xlsx-Endung hat
        if not filename.endswith('.xlsx'):
            raise ValueError(f"Die Datei '{filename}' hat keine gültige '.xlsx'-Endung.")

        # Verzeichnisprüfung
        dir_name = os.path.dirname(filename)
        if dir_name and not os.path.exists(dir_name):
            raise FileNotFoundError(f"Das Verzeichnis '{dir_name}' existiert nicht.")

        # Prüfung auf leeren DataFrame
        if df.empty:
            raise ValueError("Der DataFrame ist leer und kann nicht exportiert werden.")

        # Überprüfen, ob der DataFrame mindestens zwei Indexebenen hat
        if df.index.nlevels < 2:
            raise ValueError("Der DataFrame benötigt mindestens zwei Indexebenen für die Pivot-Darstellung.")

        # Versuch, die Pivot-Darstellung zu erstellen und zu exportieren
        df_pivot = df.unstack(level=-1)
        df_pivot.to_excel(filename, sheet_name=sheet_name)

        screen_and_log(
            f"Info: Pivot-Darstellung des DataFrames erfolgreich in '{filename}' exportiert.",
            logfile,
            screen=screen            
        )
        return True
        
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei oder Verzeichnis nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültiger DataFrame oder Dateiname: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Exportieren der Pivot-Darstellung: {e}", logfile, screen)
        return False

# -------------------------------
# Exportiert einen 2D Datafram mit Multiindex in eine Pivot-artige EXCEL Tablle, die Tabelle ist so strukturiert, dass sie eindeutige Zeilen- und Spaltenbeschriftungen hat 
#   das erlaubt das spätere Formatieren als Tabelle (ersetzt die Funktion "export_2D_df_to_excel_pivot")
# -------------------------------
def export_2D_df_to_excel_clean_table(df, filename, sheet_name='Sheet1', logfile=None, screen=True):
    """
    Exportiert einen 2D-MultiIndex-DataFrame in eine flache Excel-Tabelle.
    Die Zelle A1 enthält den Namen der ersten Indexebene.
    Die Spaltenüberschriften basieren nur auf der zweiten Indexebene.
    
    Parameter:
        df (DataFrame): Der zu exportierende DataFrame (MultiIndex erwartet).
        filename (str): Der Pfad und Name der Datei.
        sheet_name (str): Name des Arbeitsblatts (Standard: 'Sheet1').
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        bool: True bei erfolgreichem Export, False bei Fehler.
    """
    try:
        # Typprüfung
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Das übergebene Objekt ('df') ist kein gültiger pandas.DataFrame.")
        if not isinstance(filename, str):
            raise ValueError("Der Dateiname ('filename') muss ein String sein.")
        if logfile and not isinstance(logfile, str):
            raise ValueError("Der Logfile-Pfad ('logfile') muss ein String sein.")
        if not filename.endswith('.xlsx'):
            raise ValueError(f"Die Datei '{filename}' hat keine gültige '.xlsx'-Endung.")
        dir_name = os.path.dirname(filename)
        if dir_name and not os.path.exists(dir_name):
            raise FileNotFoundError(f"Das Verzeichnis '{dir_name}' existiert nicht.")
        # Prüfung auf leeren DataFrame
        if df.empty:
            raise ValueError("Der DataFrame ist leer und kann nicht exportiert werden.")
            
        if df.index.nlevels != 2:
            raise ValueError("Der DataFrame muss genau zwei Indexebenen enthalten.")

        # Pivotierung
        df_pivot = df.unstack(level=-1)

        # Nur die Spaltennamen der zweiten Indexebene verwenden
        df_pivot.columns = df_pivot.columns.get_level_values(-1)

        # Index zurücksetzen, erste Spalte bekommt ihren Namen
        index_name = df.index.names[0] or "Index"
        df_clean = df_pivot.reset_index()
        df_clean.columns.name = None
        df_clean.rename(columns={df_clean.columns[0]: index_name}, inplace=True)

        # Exportieren
        df_clean.to_excel(filename, sheet_name=sheet_name, index=False)

        screen_and_log(
            f"Info: Tabelle erfolgreich als '{filename}' exportiert.",
            logfile,
            screen=screen
        )
        return True

    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei oder Verzeichnis nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Eingabe: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Exportieren der Tabelle: {e}", logfile, screen)
        return False




# -------------------------------
# Dateien auf Verfügbarkeit prüfen
# -------------------------------
def files_availability_check(file_list, logfile=None, screen=True):
    """
    Prüft, ob Dateien vorhanden und verfügbar sind.

    Parameter:
        file_list (list): Liste von Dateipfaden.
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        bool: True, wenn alle Dateien verfügbar sind, False sonst.
    """
    # Input validation
    if not isinstance(file_list, (list, tuple)):
        raise ValueError("file_list muss eine Liste oder Tuple von Dateipfaden sein.")
    
    # Handle empty list
    if not file_list:
        screen_and_log("Info: Keine Dateien zur Prüfung angegeben.", logfile, screen)
        return True
    
    all_available = True
    available_count = 0
    
    for file_path in file_list:
        if not isinstance(file_path, str):
            screen_and_log(f"ERROR: Ungültiger Dateipfad: {file_path} (muss String sein).", logfile, screen)
            all_available = False
            continue
            
        if not os.path.isfile(file_path):
            screen_and_log(f"ERROR: Datei '{file_path}' nicht gefunden.", logfile, screen)
            all_available = False
        elif is_file_open_windows(file_path):
            screen_and_log(f"ERROR: Datei '{file_path}' ist gesperrt.", logfile, screen)
            all_available = False
        else:
            screen_and_log(f"Info: Datei '{file_path}' ist verfügbar.", logfile, screen)
            available_count += 1
    
    # Summary logging
    total_files = len(file_list)
    screen_and_log(f"Info: Verfügbarkeitscheck abgeschlossen: {available_count}/{total_files} Dateien verfügbar.", logfile, screen)
    
    return all_available

# ------------------------------------------------------------------------------
# Excel als Tabelle formatieren und erste Zeile (Header) in der Anzeige fixieren
# ------------------------------------------------------------------------------
def format_excel_as_table_with_freeze(filename, table_name="Table1", style_name="TableStyleMedium9", 
                                      freeze_first_row=True, logfile=None, screen=True):
    """
    Formatiert ein Arbeitsblatt in einer Excel-Datei als Tabelle und fixiert optional die erste Zeile.

    Parameter:
        filename (str): Der Pfad zur Excel-Datei.
        table_name (str): Der Name der Excel-Tabelle.
        style_name (str): Der Stil der Tabelle.
        freeze_first_row (bool): Ob die erste Zeile fixiert werden soll.
        logfile (str, optional): Pfad zu einer Logdatei.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        bool: True bei erfolgreichem Formatieren, False bei Fehler.
    """
    try:
        # Input validation
        if not isinstance(filename, str):
            raise ValueError("Der Dateiname ('filename') muss ein String sein.")
        if not isinstance(table_name, str):
            raise ValueError("Der Tabellenname ('table_name') muss ein String sein.")
        if not isinstance(style_name, str):
            raise ValueError("Der Stilname ('style_name') muss ein String sein.")
        if not isinstance(freeze_first_row, bool):
            raise ValueError("Der Parameter ('freeze_first_row') muss ein Boolean sein.")
            
        if not os.path.isfile(filename):
            raise FileNotFoundError(f"Die Datei '{filename}' wurde nicht gefunden.")
            
        workbook = load_workbook(filename)
        sheet = workbook.active
        
        # Check if sheet has data
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet['A1'].value is None:
            raise ValueError("Das Arbeitsblatt ist leer und kann nicht als Tabelle formatiert werden.")
        
        # Check if table name already exists
        existing_tables = [table.name for table in sheet._tables]
        if table_name in existing_tables:
            raise ValueError(f"Eine Tabelle mit dem Namen '{table_name}' existiert bereits.")
            
        table_ref = f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name=style_name, showFirstColumn=False, 
                                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
        
        if freeze_first_row:
            sheet.freeze_panes = "A2"
            
        workbook.save(filename)
        screen_and_log(f"Info: Datei '{filename}' erfolgreich als Tabelle formatiert.", logfile, screen)
        return True
        
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Eingabe: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Formatieren der Datei '{filename}': {e}", logfile, screen)
        return False

# -------------------------------
# Formatiert Zellen und Spaltenbreite einer existierenden EXCEL Datei
# -------------------------------


def format_excel_columns(filename, column_formats, column_widths=None, logfile=None, screen=True):
    """
    Öffnet eine Excel-Datei, formatiert die Spalten und passt deren Breite an.
    wenn nicht ausreichend formatiertungsangaben vorliegen bzw. übergeben werden, die letzte Spalteninformation für die folgenden Spalten verwendet  wird.

    Parameter:
        filename (str): Pfad zur Excel-Datei.
        column_formats (list): Liste von Formatstrings für Spalten (z. B. "DD.MM.YY", "#,##0.00").
        column_widths (list, optional): Liste von Breiten je Spalte. Wird die Liste überschritten,
                                        wird die letzte Breite wiederverwendet.
        logfile (str, optional): Optionaler Pfad zu einer Logdatei.
        screen (bool): Statusausgabe auf dem Bildschirm.
        
    Rückgabe:
        bool: True bei erfolgreichem Formatieren, False bei Fehler.
    """

    try:
        # Typprüfungen
        if not isinstance(filename, str):
            raise ValueError("Der Dateiname ('filename') muss ein String sein.")
        if not isinstance(column_formats, list) or not all(isinstance(fmt, str) for fmt in column_formats):
            raise ValueError("Die Spaltenformate ('column_formats') müssen eine Liste von Strings sein.")
        if column_widths and (not isinstance(column_widths, list) or not all(isinstance(w, (int, float)) for w in column_widths)):
            raise ValueError("Die Spaltenbreiten ('column_widths') müssen eine Liste von Zahlen sein.")
        if logfile and not isinstance(logfile, str):
            raise ValueError("Der Logfile-Pfad ('logfile') muss ein String sein.")

        if not os.path.isfile(filename):
            raise FileNotFoundError(f"Die Datei '{filename}' wurde nicht gefunden.")

        workbook = load_workbook(filename)
        sheet = workbook.active
        max_col = sheet.max_column

        for col_index in range(1, max_col + 1):
            # Ermittele gültige Format-Index
            fmt_index = min(col_index - 1, len(column_formats) - 1)
            fmt = column_formats[fmt_index]

            col_letter = sheet.cell(row=1, column=col_index).column_letter

            for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index, max_row=sheet.max_row):
                for cell in row:
                    cell.number_format = fmt

            if column_widths:
                width_index = min(col_index - 1, len(column_widths) - 1)
                sheet.column_dimensions[col_letter].width = column_widths[width_index]

        workbook.save(filename)
        screen_and_log(f"Info: Datei '{filename}' wurde erfolgreich formatiert und angepasst.", logfile, screen)
        return True

    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei '{filename}' wurde nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Parameter: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Formatieren der Datei '{filename}': {e}", logfile, screen)
        return False



# --------------------------
# Parquet-Dateien importieren
# --------------------------
def import_parquet(filename, logfile=None, screen=True):
    """
    Liest eine Parquet-Datei ein und gibt einen DataFrame zurück.

    Parameter:
        filename (str): Pfad zur Parquet-Datei.
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        pandas.DataFrame | None: Der DataFrame oder None bei Fehlern.
    """
    try:
        if not filename.endswith('.parquet'):
            raise ValueError(f"Die Datei '{filename}' hat keine '.parquet'-Endung.")
        if not os.path.isfile(filename):
            raise FileNotFoundError(f"Die Datei '{filename}' wurde nicht gefunden.")
        df = pd.read_parquet(filename)
        screen_and_log(f"Info: Parquet-Datei '{filename}' erfolgreich eingelesen.", logfile, screen)
        return df
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Importieren der Datei '{filename}': {e}", logfile, screen)
        return None

# ----------------------------------------------
# Funktion: Prüfen, ob eine Datei gesperrt ist
# ----------------------------------------------
def is_file_open_windows(file_path):
    """
    Überprüft, ob eine Datei unter Windows gesperrt ist.

    Parameter:
        file_path (str): Pfad der Datei.

    Rückgabe:
        bool: True, wenn die Datei gesperrt ist, False sonst.
    """
    if os.name != 'nt':
        raise OSError("Diese Funktion unterstützt nur Windows.")
    if not isinstance(file_path, str):
        raise ValueError("file_path muss ein String sein.")
    if not os.path.exists(file_path):
        return False  # Datei nicht vorhanden

    try:
        with open(file_path, 'r+b') as file:
            try:
                msvcrt.locking(file.fileno(), msvcrt.LK_NBLCK, 1)
                msvcrt.locking(file.fileno(), msvcrt.LK_UNLCK, 1)
                return False
            except IOError:
                return True
    except Exception as e:
        raise RuntimeError(f"Ein unerwarteter Fehler ist aufgetreten: {e}")

# -------------------------------
# Zentralisierte Protokollierung
# -------------------------------
def screen_and_log(message, logfile=None, screen=True, auto_log=False):
    """
    Gibt die Nachricht auf dem Bildschirm aus (je nach Bedingungen) und schreibt sie optional in ein Logfile.

    Letzte Änderung: 03.01.25

    Parameter:
        message (str): Die Nachricht, die verarbeitet werden soll.
        logfile (str, optional): Der Pfad zum Logfile. Wenn None und auto_log=True, wird automatisch 'scriptname.log' verwendet.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden sollen.
        auto_log (bool): Wenn True und logfile=None, wird automatisch ein Logfile basierend auf dem Skriptnamen erstellt.

    Rückgabe:
        str: Die formatierte Nachricht (nützlich für Tests).
    """
    if not isinstance(message, str):
        raise ValueError("Die Nachricht ('message') muss ein String sein.")
    if logfile is not None and not isinstance(logfile, str):
        raise ValueError("Der Logfile-Pfad ('logfile') muss ein String sein, wenn angegeben.")
    
    # Auto-Logfile erstellen wenn gewünscht und kein Logfile angegeben
    if logfile is None and auto_log:
        # Ermittle den Namen des ursprünglichen Skripts (nicht dieser Bibliothek)
        stack = inspect.stack()
        script_filename = None
        for frame_info in reversed(stack):  # Von außen nach innen suchen
            filename = frame_info.filename
            if not filename.endswith('Standardfunktionen_aktuell.py'):
                script_filename = filename
                break
        
        if script_filename:
            script_name = os.path.splitext(os.path.basename(script_filename))[0]
            logfile = f"{script_name}.log"
    
    # Funktionsname des Aufrufers ermitteln
    caller_function = inspect.stack()[1].function
    
    # Zeitstempel ergänzen
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Nachricht formatieren
    formatted_message = f"{current_time} - {message} (Caller Funktion: {caller_function})"

    # Bildschirm-Ausgabe steuern
    if message.strip().upper().startswith(("ERROR", "WARNING")):
        # Fall 1: Immer formatierte Nachricht anzeigen (formatierte Ausgabe)
        print(formatted_message)
    elif screen is True:
        # Fall 2a: Originale Nachricht anzeigen, wenn `screen=True`
        print(message)
    # Fall 2b: Keine Anzeige, screen=False und keine "ERROR"/"WARNING"


    # Nachricht ins Logfile schreiben, falls ein Logfile angegeben ist
    if logfile:
        try:
            log_dir = os.path.dirname(logfile) or os.getcwd()
            os.makedirs(log_dir, exist_ok=True)  # Verzeichnis erstellen, falls es nicht existiert
            with open(logfile, 'a', encoding='utf-8') as log_file:
                log_file.write(formatted_message + "\n")
        except Exception as e:
            print(f"ERROR: Konnte nicht ins Logfile '{logfile}' schreiben: {e}")
    
    return formatted_message  # Optional: Rückgabe für Tests

# -------------------------------
# Arbeitsverzeichnis setzen
# -------------------------------
def set_working_directory(path="default", logfile=None, screen=True):
    """
    Setzt das aktuelle Arbeitsverzeichnis.

    Parameter:
        path (str): 
            - Wenn "default", wird das Arbeitsverzeichnis auf das Verzeichnis gesetzt, 
              von dem aus das Skript gestartet wurde.
            - Andernfalls wird der übergebene Pfad als Arbeitsverzeichnis verwendet.
        logfile (str, optional): Der Pfad zum Logfile. Wenn None, wird keine Protokollierung ins Logfile durchgeführt.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        None

    Hinweise:
        - Falls der Pfad ungültig ist, wird eine Fehlermeldung ausgegeben und das Verzeichnis nicht geändert.
        - Änderungen des Arbeitsverzeichnisses werden protokolliert.
    """
    try:
        # Typprüfung des Eingabeparameters
        if not isinstance(path, str):
            raise ValueError("Der Pfad ('path') muss ein String sein.")

        # Setzt das Arbeitsverzeichnis auf das Verzeichnis, aus dem das Caller Skript gestartet wurde
        if path == "default":
            # Ermitteln des Verzeichnisses des Callers
            caller_frame = inspect.stack()[1]
            caller_filename = caller_frame.filename
            caller_directory = os.path.dirname(os.path.abspath(caller_filename))
            os.chdir(caller_directory)
            screen_and_log(f"Info: Arbeitsverzeichnis wurde auf das Verzeichnis des Callers gesetzt: {caller_directory}", logfile, screen)
        else:
            # Prüft, ob der angegebene Pfad existiert und ein Verzeichnis ist
            if os.path.exists(path) and os.path.isdir(path):
                os.chdir(path)
                screen_and_log(f"Info: Arbeitsverzeichnis wurde auf '{path}' gesetzt.", logfile, screen)
            else:
                raise FileNotFoundError(f"Der Pfad '{path}' ist nicht verfügbar oder kein gültiges Verzeichnis.")

    except PermissionError:
        screen_and_log(f"ERROR: Keine Berechtigung, das Arbeitsverzeichnis auf '{path}' zu setzen.", logfile, screen)
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Der angegebene Pfad ist ungültig: {e}", logfile, screen)
    except Exception as e:
        screen_and_log(f"ERROR: Ein unerwarteter Fehler ist aufgetreten: {e}", logfile, screen)

# -------------------------------
# Einstellungen aus INI-Datei laden
# -------------------------------
import os
import configparser
import ast  # Ermöglicht sicheres Parsen von Python-Literalen (z. B. dict, list, int)

def settings_import(file_name):
    """
    Liest eine INI-Datei ein und gibt die Inhalte als Dictionary zurück.
    Unterstützt strukturierte Werte wie Dictionaries in einer Zeile (z. B. für Export-Optionen).

    Formatbeispiel in der INI:
        [Export]
        values_month_to_excel = {"enabled": true, "filename": "file.xlsx", "column_formats": ["DD.MM.YY"], "column_widths": [12]}

    Rückgabe:
        dict: Strukturierte Einstellungen, z. B. settings["Export"]["values_month_to_excel"]["filename"]
    """
    
    try:
        # Input validation
        if not isinstance(file_name, str):
            raise ValueError("Der Dateiname ('file_name') muss ein String sein.")
            
        # Prüfe, ob die Datei existiert
        if not os.path.isfile(file_name):
            raise FileNotFoundError(f"Die Datei '{file_name}' wurde nicht gefunden.")

        # Lade die INI-Datei
        config = configparser.ConfigParser(interpolation=None)
        config.read(file_name)

        # Zieldatenstruktur: Dictionary mit Abschnittsnamen (z. B. "Export") als Schlüssel
        settings = {}

        # Schleife über alle Abschnitte der INI-Datei
        for section in config.sections():
            settings[section] = {}  # Neuen Abschnitt vorbereiten

            # Schleife über alle Schlüssel/Wert-Paare in diesem Abschnitt
            for key, value in config.items(section):
                value = value.strip()  # Whitespace am Anfang/Ende entfernen
                
                # Versuche, strukturierte Werte (Dictionaries, Listen etc.) als Python-Objekt zu parsen
                if value.startswith("{") and value.endswith("}"):
                    try:
                        # Convert JSON-style boolean literals to Python literals for ast.literal_eval
                        python_value = value.replace(' true', ' True').replace(' false', ' False')
                        python_value = python_value.replace(':true', ':True').replace(':false', ':False')
                        python_value = python_value.replace('[true', '[True').replace('[false', '[False')
                        python_value = python_value.replace(',true', ',True').replace(',false', ',False')
                        python_value = python_value.replace('{true', '{True').replace('{false', '{False')

                        parsed = ast.literal_eval(python_value)  # Sicheres Parsen von Python-Datentypen
                        settings[section][key] = parsed   # Im Dictionary speichern
                        continue  # Weiter zum nächsten Eintrag
                    except Exception as e:
                        # Parsing fehlgeschlagen → als Fallback weitermachen
                        screen_and_log(f"WARNING: Kann Wert für '{section}:{key}' nicht als Dictionary parsen: {e}",
                                     None, True, auto_log=True)

                # Bool-Werte erkennen und umwandeln
                if value.lower() in ['true', 'false']:
                    settings[section][key] = value.lower() == 'true'
                # Ganzzahl oder Kommazahl erkennen
                elif value.replace('.', '', 1).isdigit():
                    settings[section][key] = float(value) if '.' in value else int(value)
                # Kommagetrennte Liste → z. B. "A,B,C"
                elif ',' in value:
                    settings[section][key] = [v.strip() for v in value.split(',')]
                else:
                    # Fallback: einfacher String
                    settings[section][key] = value

        return settings

    except Exception as e:
        # Allgemeine Fehlerbehandlung (z. B. Datei beschädigt)
        screen_and_log(f"ERROR: Fehler beim Laden der Einstellungen: {e}", None, True, auto_log=True)
        return None

# -------------------------------
# Enhanced ConfigParser with Structured Data Support
# -------------------------------
class StructuredConfigParser(configparser.ConfigParser):
    """
    Enhanced ConfigParser with automatic type conversion and settings_import compatibility.

    Features:
    - All standard ConfigParser methods work unchanged
    - Automatic type conversion (bool, int, float, list, dict)
    - settings_import() style dict export with to_dict()
    - Structured data support like {"key": "value"} in INI files
    - Same parsing logic as settings_import() function

    Usage:
        config = StructuredConfigParser()
        config.read('config.ini')

        # Traditional access (unchanged)
        value = config.get('Section', 'key', fallback='default')

        # Enhanced access with type conversion
        typed_value = config.get_structured('Section', 'key', fallback=default)

        # Export as dict (like settings_import)
        settings_dict = config.to_dict()
    """

    def __init__(self, *args, **kwargs):
        """Initialize with interpolation disabled (like settings_import)."""
        kwargs.setdefault('interpolation', None)
        super().__init__(*args, **kwargs)

    def get_structured(self, section, option, fallback=None):
        """
        Get configuration value with automatic type conversion.

        Uses same parsing logic as settings_import() function.

        Args:
            section (str): Section name in INI file
            option (str): Option name in section
            fallback: Default value if option not found

        Returns:
            Parsed value with appropriate Python type (dict, list, bool, int, float, str)
        """
        try:
            value = self.get(section, option).strip()
            return self._parse_value(value)
        except (configparser.NoSectionError, configparser.NoOptionError):
            return fallback

    def _parse_value(self, value):
        """
        Parse string value to appropriate Python type.

        Uses identical logic to settings_import() function for consistency.

        Args:
            value (str): Raw string value from INI file

        Returns:
            Parsed value (dict, list, bool, int, float, or str)
        """
        # Try structured data (dict/list/tuple) - same as settings_import
        if ((value.startswith("{") and value.endswith("}")) or
            (value.startswith("[") and value.endswith("]")) or
            (value.startswith("(") and value.endswith(")"))):
            try:
                # Convert JSON-style boolean literals to Python literals for ast.literal_eval
                python_value = value.replace(' true', ' True').replace(' false', ' False')
                python_value = python_value.replace(':true', ':True').replace(':false', ':False')
                python_value = python_value.replace('[true', '[True').replace('[false', '[False')
                python_value = python_value.replace(',true', ',True').replace(',false', ',False')
                python_value = python_value.replace('{true', '{True').replace('{false', '{False')

                parsed = ast.literal_eval(python_value)  # Same as settings_import
                return parsed
            except Exception:
                # Parsing failed → continue with other patterns
                pass

        # Bool-Werte erkennen und umwandeln - same as settings_import
        if value.lower() in ['true', 'false']:
            return value.lower() == 'true'

        # Ganzzahl oder Kommazahl erkennen - same as settings_import
        elif value.replace('.', '', 1).isdigit():
            return float(value) if '.' in value else int(value)

        # Kommagetrennte Liste → z. B. "A,B,C" - same as settings_import
        # But skip if it looks like it might be structured data that failed to parse
        elif ',' in value and not ('{' in value or '[' in value or '(' in value):
            return [v.strip() for v in value.split(',')]

        else:
            # Fallback: einfacher String - same as settings_import
            return value

    def to_dict(self):
        """
        Export entire configuration as nested dictionary.

        Returns same format as settings_import() function:
        {
            "Section1": {"key1": parsed_value1, "key2": parsed_value2},
            "Section2": {"key3": parsed_value3}
        }

        Returns:
            dict: Nested dictionary with parsed values
        """
        settings = {}
        for section_name in self.sections():
            settings[section_name] = {}
            for key, value in self.items(section_name):
                settings[section_name][key] = self._parse_value(value)
        return settings

    def get_section_dict(self, section):
        """
        Get entire section as dictionary with parsed values.

        Args:
            section (str): Section name

        Returns:
            dict: All keys in section with parsed values, empty dict if section not found
        """
        if not self.has_section(section):
            return {}

        section_dict = {}
        for key, value in self.items(section):
            section_dict[key] = self._parse_value(value)
        return section_dict

def load_structured_config(file_name, logfile=None, screen=True):
    """
    Load INI file with structured data support and comprehensive error handling.

    Enhanced version of settings_import() that returns StructuredConfigParser object
    instead of dict, allowing both traditional ConfigParser usage and structured data access.

    Args:
        file_name (str): Path to INI configuration file
        logfile (str, optional): Path to log file for error logging
        screen (bool): Whether to show messages on screen

    Returns:
        StructuredConfigParser: Enhanced config parser with type conversion
        None: If loading failed (with error logged)

    Raises:
        ValueError: If file_name is not a string
        FileNotFoundError: If configuration file not found

    Usage:
        # Method 1: Use as enhanced ConfigParser
        config = load_structured_config("config.ini")
        if config:
            value = config.get_structured("Section", "key", default_value)
            section_data = config.get_section_dict("Section")

        # Method 2: Export as dict (like settings_import)
        config = load_structured_config("config.ini")
        if config:
            settings = config.to_dict()  # Same format as settings_import()
    """
    try:
        # Input validation (same as settings_import)
        if not isinstance(file_name, str):
            raise ValueError("Der Dateiname ('file_name') muss ein String sein.")

        # Check if file exists (same as settings_import)
        if not os.path.isfile(file_name):
            raise FileNotFoundError(f"Die Datei '{file_name}' wurde nicht gefunden.")

        # Load INI file with structured parser
        config = StructuredConfigParser()
        config.read(file_name, encoding='utf-8')  # Use UTF-8 for better compatibility

        screen_and_log(f"Info: Konfigurationsdatei '{file_name}' erfolgreich geladen.", logfile, screen)
        return config

    except Exception as e:
        # Same error handling as settings_import
        screen_and_log(f"ERROR: Fehler beim Laden der Konfiguration: {e}", logfile, screen, auto_log=True)
        return None

def settings_import_structured(file_name, logfile=None, screen=True):
    """
    Enhanced version of settings_import() returning StructuredConfigParser.

    Provides both dict export (backwards compatible) and ConfigParser object access.

    Args:
        file_name (str): Path to INI configuration file
        logfile (str, optional): Path to log file
        screen (bool): Whether to show messages on screen

    Returns:
        StructuredConfigParser or None: Enhanced config parser or None if failed

    Usage:
        # Backwards compatible with settings_import
        config = settings_import_structured("config.ini")
        if config:
            settings = config.to_dict()  # Same as original settings_import()

        # Plus enhanced ConfigParser features
        if config:
            value = config.get_structured("Section", "key", fallback)
    """
    return load_structured_config(file_name, logfile, screen)

def load_structured_config_with_validation(file_name, validation_error_class=None):
    """
    Load INI file with structured data support for account statements scripts.

    This function is specifically designed for account statements scripts that use
    ValidationError for error handling instead of returning None.

    Args:
        file_name (str): Path to INI configuration file
        validation_error_class: ValidationError class to use for exceptions
                               (defaults to creating a simple Exception subclass)

    Returns:
        StructuredConfigParser: Enhanced config parser with type conversion

    Raises:
        ValidationError: If file not found or cannot be read

    Usage:
        # In account statements scripts
        try:
            from core.types import ValidationError
        except ImportError:
            class ValidationError(Exception):
                pass

        config = load_structured_config_with_validation("as.ini", ValidationError)
        value = config.get_structured("Section", "key", default)
    """
    # Create default ValidationError if none provided
    if validation_error_class is None:
        class ValidationError(Exception):
            pass
        validation_error_class = ValidationError

    # Input validation
    if not isinstance(file_name, str):
        raise validation_error_class("Der Dateiname ('file_name') muss ein String sein.")

    file_path = Path(file_name)
    if not file_path.exists():
        raise validation_error_class(f"Configuration file not found: {file_path}")

    # Load with structured parser
    config = StructuredConfigParser()
    try:
        config.read(file_path, encoding='utf-8')
    except Exception as e:
        raise validation_error_class(f"Error reading configuration file: {e}")

    return config


# -------------------------------
# Erweiterte Protokollierung (Advanced Logging)
# -------------------------------

import time
from typing import Dict, List, Any, Optional
from collections import Counter, defaultdict
from dataclasses import dataclass, field


@dataclass
class VerarbeitungsMetriken:
    """Detaillierte Metriken für Verarbeitungsoperationen."""
    dateien_verarbeitet: int = 0
    dateien_übersprungen: int = 0
    zeilen_nach_blatt: Dict[str, int] = field(default_factory=dict)
    fehler_nach_typ: Counter = field(default_factory=Counter)
    fehler_nach_datei: Dict[str, List[str]] = field(default_factory=lambda: defaultdict(list))
    zeilen_gesamt_hinzugefügt: int = 0
    doppelte_zeilen_übersprungen: int = 0
    verarbeitungszeit_sekunden: float = 0.0
    startzeit: Optional[float] = None

    def starte_zeitmessung(self) -> None:
        """Starte die Zeitmessung für die Verarbeitungsoperation."""
        self.startzeit = time.time()

    def stoppe_zeitmessung(self) -> None:
        """Stoppe die Zeitmessung und erfasse die Dauer."""
        if self.startzeit:
            self.verarbeitungszeit_sekunden = time.time() - self.startzeit

    def füge_blatt_zeilen_hinzu(self, blatt_name: str, zeilen_anzahl: int) -> None:
        """Füge Zeilenzahl für ein Blatt hinzu."""
        self.zeilen_nach_blatt[blatt_name] = self.zeilen_nach_blatt.get(blatt_name, 0) + zeilen_anzahl
        self.zeilen_gesamt_hinzugefügt += zeilen_anzahl

    def füge_fehler_hinzu(self, fehler_typ: str, dateiname: str, fehler_nachricht: str) -> None:
        """Erfasse einen Fehler."""
        self.fehler_nach_typ[fehler_typ] += 1
        self.fehler_nach_datei[dateiname].append(f"{fehler_typ}: {fehler_nachricht}")

    def erfasse_datei_verarbeitet(self) -> None:
        """Erfasse, dass eine Datei erfolgreich verarbeitet wurde."""
        self.dateien_verarbeitet += 1

    def erfasse_datei_übersprungen(self, grund: str = "") -> None:
        """Erfasse, dass eine Datei übersprungen wurde."""
        self.dateien_übersprungen += 1
        if grund:
            self.fehler_nach_typ[f"übersprungen: {grund}"] += 1


class ErweiterterLogger:
    """
    Erweiterte Protokollierungsklasse für komplexe Verarbeitungsabläufe.

    Bietet strukturierte Protokollierung mit Metriken-Erfassung, Performance-Tracking
    und detaillierte Fehlerbehandlung. Alternative zu screen_and_log() für komplexe
    Skripte, die folgende Features benötigen:
    - Strukturierte Log-Level (INFO/WARNING/ERROR/DEBUG)
    - Verarbeitungsmetriken und Statistiken
    - Performance-Tracking
    - Batch-Operationen und Zusammenfassungen

    Eigenschaften:
    - Doppelte Ausgabe: Datei-Protokollierung + optionale Konsolen-Ausgabe
    - Standardisierte Nachrichtenformatierung
    - Verarbeitungsmetriken und Performance-Tracking
    - Fehlerbehandlung und Fortschrittsberichte
    """

    def __init__(self, protokolldatei: str, bildschirm_ausgabe: bool = True, skript_name: str = ""):
        """
        Initialisiere Logger mit konsistenter Konfiguration.

        Args:
            protokolldatei: Pfad zur Protokolldatei
            bildschirm_ausgabe: Konsolen-Ausgabe aktivieren
            skript_name: Skript-Identifikator für Protokollnachrichten
        """
        self.protokolldatei = Path(protokolldatei)
        self.bildschirm_ausgabe = bildschirm_ausgabe
        self.skript_name = skript_name
        self.metriken = VerarbeitungsMetriken()

        # Stelle sicher, dass das Protokoll-Verzeichnis existiert
        self.protokolldatei.parent.mkdir(parents=True, exist_ok=True)

    def protokolliere(self, nachricht: str, level: str = "INFO") -> None:
        """Protokolliere eine Nachricht in Datei und optional Konsole."""
        zeitstempel = time.strftime("%Y-%m-%d %H:%M:%S")

        # Dateiformat: [zeitstempel] skript_name LEVEL: nachricht
        skript_teil = f" {self.skript_name}" if self.skript_name else ""
        datei_formatierte_nachricht = f"[{zeitstempel}]{skript_teil} {level}: {nachricht}"

        # Bildschirmformat: [zeitstempel] LEVEL: nachricht
        bildschirm_formatierte_nachricht = f"[{zeitstempel}] {level}: {nachricht}"

        # Konsolen-Ausgabe
        if self.bildschirm_ausgabe:
            try:
                print(bildschirm_formatierte_nachricht)
            except Exception:
                pass  # Fehler bei Konsolen-Ausgabe nicht weiterleiten

        # Datei-Ausgabe
        try:
            with open(self.protokolldatei, "a", encoding="utf-8") as f:
                f.write(datei_formatierte_nachricht + "\n")
        except Exception:
            pass  # Fehler bei Protokollierung nicht weiterleiten

    def info(self, nachricht: str) -> None:
        """Protokolliere eine Info-Nachricht."""
        self.protokolliere(nachricht, "INFO")

    def warnung(self, nachricht: str) -> None:
        """Protokolliere eine Warnung."""
        self.protokolliere(nachricht, "WARNING")

    def fehler(self, nachricht: str) -> None:
        """Protokolliere einen Fehler."""
        self.protokolliere(nachricht, "ERROR")

    def debug(self, nachricht: str) -> None:
        """Protokolliere eine Debug-Nachricht."""
        self.protokolliere(nachricht, "DEBUG")

    def starte_verarbeitung(self) -> None:
        """Markiere den Start der Verarbeitung."""
        self.metriken.starte_zeitmessung()
        self.info("Starte Account-Statements-Verarbeitung")

    def beende_verarbeitung(self) -> None:
        """Markiere das Ende der Verarbeitung und protokolliere Zusammenfassung."""
        self.metriken.stoppe_zeitmessung()
        self.protokolliere_zusammenfassung()

    def erfasse_datei_verarbeitet(self, dateiname: str, blatt_name: str, zeilen_anzahl: int, metadaten: Dict[str, Any]) -> None:
        """Erfasse erfolgreiche Datei-Verarbeitung."""
        self.metriken.erfasse_datei_verarbeitet()
        self.metriken.füge_blatt_zeilen_hinzu(blatt_name, zeilen_anzahl)

        parser_info = f"{metadaten.get('parser', 'unbekannt')}"
        erkennungs_info = f"{metadaten.get('detected_via', [])}"

        self.info(f"Verarbeitet {dateiname} -> {blatt_name}: +{zeilen_anzahl} Zeilen via {parser_info} {erkennungs_info}")

    def erfasse_datei_übersprungen(self, dateiname: str, grund: str) -> None:
        """Erfasse übersprungene Datei mit Grund."""
        self.metriken.erfasse_datei_übersprungen(grund)
        self.warnung(f"Übersprungen {dateiname}: {grund}")

    def erfasse_fehler(self, dateiname: str, fehler_typ: str, fehler_nachricht: str) -> None:
        """Erfasse das Auftreten eines Fehlers."""
        self.metriken.füge_fehler_hinzu(fehler_typ, dateiname, fehler_nachricht)
        self.fehler(f"Fehler in {dateiname} ({fehler_typ}): {fehler_nachricht}")

    def erfasse_archiv_erfolg(self, dateiname: str, ziel: str) -> None:
        """Erfasse erfolgreiche Datei-Archivierung."""
        self.info(f"Archiviert: {dateiname} -> {ziel}")

    def erfasse_archiv_fehler(self, dateiname: str, fehler: str) -> None:
        """Erfasse Archivierungsfehler."""
        self.fehler(f"Archivierungsfehler für {dateiname}: {fehler}")

    def protokolliere_fortschritt(self, aktuell: int, gesamt: int, element_name: str = "Dateien") -> None:
        """Protokolliere Fortschrittsinformationen."""
        if gesamt > 0:
            prozent = (aktuell / gesamt) * 100
            self.info(f"Fortschritt: {aktuell}/{gesamt} {element_name} ({prozent:.1f}%)")

    def protokolliere_duplikat_ergebnisse(self, blatt_name: str, gesamt_neu: int, bereits_vorhanden: int, wird_hinzugefügt: int) -> None:
        """Protokolliere Duplikat-Ergebnisse für ein Blatt."""
        self.info(f"{blatt_name}: gesamt_neu={gesamt_neu}, bereits_vorhanden={bereits_vorhanden}, wird_hinzugefügt={wird_hinzugefügt}")
        self.metriken.doppelte_zeilen_übersprungen += bereits_vorhanden

    def protokolliere_zusammenfassung(self) -> None:
        """Protokolliere Verarbeitungszusammenfassung."""
        self.info("===== Verarbeitungszusammenfassung =====")
        self.info(f"Dateien verarbeitet: {self.metriken.dateien_verarbeitet}")
        self.info(f"Dateien übersprungen: {self.metriken.dateien_übersprungen}")
        self.info(f"Zeilen gesamt hinzugefügt: {self.metriken.zeilen_gesamt_hinzugefügt}")
        self.info(f"Doppelte Zeilen übersprungen: {self.metriken.doppelte_zeilen_übersprungen}")
        self.info(f"Verarbeitungszeit: {self.metriken.verarbeitungszeit_sekunden:.2f} Sekunden")

        if self.metriken.zeilen_nach_blatt:
            self.info("Zeilen nach Blatt:")
            for blatt, anzahl in self.metriken.zeilen_nach_blatt.items():
                self.info(f"  {blatt}: {anzahl}")

        if self.metriken.fehler_nach_typ:
            self.info("Aufgetretene Fehler:")
            for fehler_typ, anzahl in self.metriken.fehler_nach_typ.items():
                self.info(f"  {fehler_typ}: {anzahl}")

        # Performance-Info
        if self.metriken.verarbeitungszeit_sekunden > 0:
            dateien_pro_sekunde = self.metriken.dateien_verarbeitet / self.metriken.verarbeitungszeit_sekunden
            zeilen_pro_sekunde = self.metriken.zeilen_gesamt_hinzugefügt / self.metriken.verarbeitungszeit_sekunden
            self.info(f"Performance: {dateien_pro_sekunde:.2f} Dateien/Sek, {zeilen_pro_sekunde:.1f} Zeilen/Sek")

    def hole_metriken(self) -> VerarbeitungsMetriken:
        """Hole das aktuelle Metriken-Objekt."""
        return self.metriken


def erstelle_erweiterten_logger(protokolldatei_pfad: str, bildschirm_ausgabe: bool = True,
                               skript_name: str = "") -> ErweiterterLogger:
    """
    Erstelle eine ErweiterterLogger-Instanz.

    Alternative zu screen_and_log() für Skripte, die folgende Features benötigen:
    - Strukturierte Protokoll-Level (INFO/WARNING/ERROR/DEBUG)
    - Verarbeitungsmetriken und Statistiken
    - Performance-Tracking
    - Batch-Operationen und detaillierte Zusammenfassungen

    Args:
        protokolldatei_pfad: Pfad zur Protokolldatei
        bildschirm_ausgabe: Konsolen-Ausgabe aktivieren (Standard: True)
        skript_name: Name des aufrufenden Skripts (z.B. 'as_1_extract')

    Returns:
        ErweiterterLogger: Konfigurierte Logger-Instanz

    Verwendung:
        from ahlib import erstelle_erweiterten_logger
        logger = erstelle_erweiterten_logger("verarbeitung.log", True, "mein_skript")
        logger.info("Verarbeitung gestartet")
        logger.erfasse_datei_verarbeitet("datei.csv", "Blatt1", 1000, {"parser": "CSV"})
        logger.protokolliere_zusammenfassung()
    """
    return ErweiterterLogger(protokolldatei_pfad, bildschirm_ausgabe, skript_name)
