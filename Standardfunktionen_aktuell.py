import os
import pandas as pd
import msvcrt
from datetime import datetime
import inspect
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import configparser


# --------------------------
# Parquet-Dateien exportieren
# --------------------------
def export_df_to_parquet(df, filename, compression=None, logfile=None, screen=True):
    """
    Exportiert einen DataFrame in eine Parquet-Datei.

    Parameter:
        df (pandas.DataFrame): Der zu exportierende DataFrame.
        filename (str): Der Pfad zur Zieldatei.
        compression (str, optional): Komprimierung ('snappy', 'gzip', 'brotli', None).
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.

    Rückgabe:
        bool: True bei erfolgreichem Export, False bei Fehler.
    """
    try:
        # Typprüfung DataFrame
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Das übergebene Objekt ('df') ist kein gültiger pandas.DataFrame.")
        
        # Prüfung auf leeren DataFrame
        if df.empty:
            raise ValueError("Der DataFrame ist leer und kann nicht exportiert werden.")
        
        # Dateiendung prüfen
        if not filename.endswith('.parquet'):
            raise ValueError(f"Die Datei '{filename}' hat keine '.parquet'-Endung.")
        
        # Verzeichnis erstellen falls nicht vorhanden
        dir_name = os.path.dirname(filename)
        if dir_name and not os.path.exists(dir_name):
            os.makedirs(dir_name, exist_ok=True)
            screen_and_log(f"Info: Verzeichnis '{dir_name}' wurde erstellt.", logfile, screen)
        
        # DataFrame exportieren
        df.to_parquet(filename, compression=compression)
        screen_and_log(f"Info: DataFrame erfolgreich in '{filename}' exportiert.", logfile, screen)
        return True
        
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei oder Verzeichnis nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Eingabe: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Exportieren des DataFrames in '{filename}': {e}", logfile, screen)
        return False

# --------------------------
# Excel-Dateien exportieren
# --------------------------
def export_df_to_excel(df, filename, sheet_name='Sheet1', logfile=None, screen=True):
    """
    Exportiert einen DataFrame in eine Excel-Datei.

    Parameter:
        df (pandas.DataFrame): Der zu exportierende DataFrame.
        filename (str): Der Pfad zur Zieldatei.
        sheet_name (str): Name des Arbeitsblatts (Standard: 'Sheet1').
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.

    Rückgabe:
        bool: True bei erfolgreichem Export, False bei Fehler.
    """
    try:
        # Typprüfung DataFrame
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Das übergebene Objekt ('df') ist kein gültiger pandas.DataFrame.")
        
        # Prüfung auf leeren DataFrame
        if df.empty:
            raise ValueError("Der DataFrame ist leer und kann nicht exportiert werden.")
        
        # Dateiendung prüfen
        if not filename.endswith('.xlsx'):
            raise ValueError(f"Die Datei '{filename}' hat keine '.xlsx'-Endung.")
        
        # Verzeichnis erstellen falls nicht vorhanden
        dir_name = os.path.dirname(filename)
        if dir_name and not os.path.exists(dir_name):
            os.makedirs(dir_name, exist_ok=True)
            screen_and_log(f"Info: Verzeichnis '{dir_name}' wurde erstellt.", logfile, screen)
        
        # DataFrame exportieren - prüfe ob Index als Spalte benötigt wird
        if df.index.name or (hasattr(df.index, 'names') and any(df.index.names)):
            # Index hat einen Namen (z.B. 'date') -> als Spalte exportieren
            df_export = df.reset_index()
            df_export.to_excel(filename, sheet_name=sheet_name, index=False)
        else:
            # Index ist Standard-Index -> ohne Index exportieren
            df.to_excel(filename, sheet_name=sheet_name, index=False)
        screen_and_log(f"Info: DataFrame erfolgreich in '{filename}' exportiert.", logfile, screen)
        return True
        
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei oder Verzeichnis nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Eingabe: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Exportieren des DataFrames in '{filename}': {e}", logfile, screen)
        return False

# -------------------------------
# Exportiert einen 2D Datafram mit Multiindex in eine Pivot-artige EXCEL Tablle 
# -------------------------------
def export_2D_df_to_excel_pivot(df, filename, sheet_name='Sheet1', logfile=None, screen=True):
    """
    Exportiert den übergebenen DataFrame in eine Pivot-Darstellung als Excel-Datei.
    Zeilen enthalten die erste Indexebene, Spalten die zweite Indexebene und die Zellen den Wert.

    Parameter:
        df (DataFrame): Der zu exportierende DataFrame (MultiIndex erwartet).
        filename (str): Der Pfad und Name der Datei, in die der DataFrame exportiert werden soll.
        sheet_name (str): Name des Arbeitsblatts (Standard: 'Sheet1').
        logfile (str, optional): Der Pfad zum Logfile. Wenn None, wird keine Protokollierung ins Logfile durchgeführt.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.

    Rückgabe:
        bool: True bei erfolgreichem Export, False bei Fehler.

    Hinweise:
        - Der DataFrame muss mindestens zwei Indexebenen enthalten.
        - Der Dateiname muss auf ".xlsx" enden.
        - Das Zielverzeichnis muss existieren und beschreibbar sein.
        - Wenn der Export fehlschlägt, wird eine Fehlermeldung protokolliert.
    """
    try:
        # Typprüfung der Eingabeparameter
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Das übergebene Objekt ('df') ist kein gültiger pandas.DataFrame.")
        if not isinstance(filename, str):
            raise ValueError("Der Dateiname ('filename') muss ein String sein.")
        if logfile and not isinstance(logfile, str):
            raise ValueError("Der Logfile-Pfad ('logfile') muss ein String sein.")

        # Überprüfen, ob der Dateiname eine .xlsx-Endung hat
        if not filename.endswith('.xlsx'):
            raise ValueError(f"Die Datei '{filename}' hat keine gültige '.xlsx'-Endung.")

        # Verzeichnisprüfung
        dir_name = os.path.dirname(filename)
        if dir_name and not os.path.exists(dir_name):
            raise FileNotFoundError(f"Das Verzeichnis '{dir_name}' existiert nicht.")

        # Prüfung auf leeren DataFrame
        if df.empty:
            raise ValueError("Der DataFrame ist leer und kann nicht exportiert werden.")

        # Überprüfen, ob der DataFrame mindestens zwei Indexebenen hat
        if df.index.nlevels < 2:
            raise ValueError("Der DataFrame benötigt mindestens zwei Indexebenen für die Pivot-Darstellung.")

        # Versuch, die Pivot-Darstellung zu erstellen und zu exportieren
        df_pivot = df.unstack(level=-1)
        df_pivot.to_excel(filename, sheet_name=sheet_name)

        screen_and_log(
            f"Info: Pivot-Darstellung des DataFrames erfolgreich in '{filename}' exportiert.",
            logfile,
            screen=screen            
        )
        return True
        
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei oder Verzeichnis nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültiger DataFrame oder Dateiname: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Exportieren der Pivot-Darstellung: {e}", logfile, screen)
        return False

# -------------------------------
# Exportiert einen 2D Datafram mit Multiindex in eine Pivot-artige EXCEL Tablle, die Tabelle ist so strukturiert, dass sie eindeutige Zeilen- und Spaltenbeschriftungen hat 
#   das erlaubt das spätere Formatieren als Tabelle (ersetzt die Funktion "export_2D_df_to_excel_pivot")
# -------------------------------
def export_2D_df_to_excel_clean_table(df, filename, sheet_name='Sheet1', logfile=None, screen=True):
    """
    Exportiert einen 2D-MultiIndex-DataFrame in eine flache Excel-Tabelle.
    Die Zelle A1 enthält den Namen der ersten Indexebene.
    Die Spaltenüberschriften basieren nur auf der zweiten Indexebene.
    
    Parameter:
        df (DataFrame): Der zu exportierende DataFrame (MultiIndex erwartet).
        filename (str): Der Pfad und Name der Datei.
        sheet_name (str): Name des Arbeitsblatts (Standard: 'Sheet1').
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        bool: True bei erfolgreichem Export, False bei Fehler.
    """
    try:
        # Typprüfung
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Das übergebene Objekt ('df') ist kein gültiger pandas.DataFrame.")
        if not isinstance(filename, str):
            raise ValueError("Der Dateiname ('filename') muss ein String sein.")
        if logfile and not isinstance(logfile, str):
            raise ValueError("Der Logfile-Pfad ('logfile') muss ein String sein.")
        if not filename.endswith('.xlsx'):
            raise ValueError(f"Die Datei '{filename}' hat keine gültige '.xlsx'-Endung.")
        dir_name = os.path.dirname(filename)
        if dir_name and not os.path.exists(dir_name):
            raise FileNotFoundError(f"Das Verzeichnis '{dir_name}' existiert nicht.")
        # Prüfung auf leeren DataFrame
        if df.empty:
            raise ValueError("Der DataFrame ist leer und kann nicht exportiert werden.")
            
        if df.index.nlevels != 2:
            raise ValueError("Der DataFrame muss genau zwei Indexebenen enthalten.")

        # Pivotierung
        df_pivot = df.unstack(level=-1)

        # Nur die Spaltennamen der zweiten Indexebene verwenden
        df_pivot.columns = df_pivot.columns.get_level_values(-1)

        # Index zurücksetzen, erste Spalte bekommt ihren Namen
        index_name = df.index.names[0] or "Index"
        df_clean = df_pivot.reset_index()
        df_clean.columns.name = None
        df_clean.rename(columns={df_clean.columns[0]: index_name}, inplace=True)

        # Exportieren
        df_clean.to_excel(filename, sheet_name=sheet_name, index=False)

        screen_and_log(
            f"Info: Tabelle erfolgreich als '{filename}' exportiert.",
            logfile,
            screen=screen
        )
        return True

    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei oder Verzeichnis nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Eingabe: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Exportieren der Tabelle: {e}", logfile, screen)
        return False




# -------------------------------
# Dateien auf Verfügbarkeit prüfen
# -------------------------------
def files_availability_check(file_list, logfile=None, screen=True):
    """
    Prüft, ob Dateien vorhanden und verfügbar sind.

    Parameter:
        file_list (list): Liste von Dateipfaden.
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        bool: True, wenn alle Dateien verfügbar sind, False sonst.
    """
    # Input validation
    if not isinstance(file_list, (list, tuple)):
        raise ValueError("file_list muss eine Liste oder Tuple von Dateipfaden sein.")
    
    # Handle empty list
    if not file_list:
        screen_and_log("Info: Keine Dateien zur Prüfung angegeben.", logfile, screen)
        return True
    
    all_available = True
    available_count = 0
    
    for file_path in file_list:
        if not isinstance(file_path, str):
            screen_and_log(f"ERROR: Ungültiger Dateipfad: {file_path} (muss String sein).", logfile, screen)
            all_available = False
            continue
            
        if not os.path.isfile(file_path):
            screen_and_log(f"ERROR: Datei '{file_path}' nicht gefunden.", logfile, screen)
            all_available = False
        elif is_file_open_windows(file_path):
            screen_and_log(f"ERROR: Datei '{file_path}' ist gesperrt.", logfile, screen)
            all_available = False
        else:
            screen_and_log(f"Info: Datei '{file_path}' ist verfügbar.", logfile, screen)
            available_count += 1
    
    # Summary logging
    total_files = len(file_list)
    screen_and_log(f"Info: Verfügbarkeitscheck abgeschlossen: {available_count}/{total_files} Dateien verfügbar.", logfile, screen)
    
    return all_available

# ------------------------------------------------------------------------------
# Excel als Tabelle formatieren und erste Zeile (Header) in der Anzeige fixieren
# ------------------------------------------------------------------------------
def format_excel_as_table_with_freeze(filename, table_name="Table1", style_name="TableStyleMedium9", 
                                      freeze_first_row=True, logfile=None, screen=True):
    """
    Formatiert ein Arbeitsblatt in einer Excel-Datei als Tabelle und fixiert optional die erste Zeile.

    Parameter:
        filename (str): Der Pfad zur Excel-Datei.
        table_name (str): Der Name der Excel-Tabelle.
        style_name (str): Der Stil der Tabelle.
        freeze_first_row (bool): Ob die erste Zeile fixiert werden soll.
        logfile (str, optional): Pfad zu einer Logdatei.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        bool: True bei erfolgreichem Formatieren, False bei Fehler.
    """
    try:
        # Input validation
        if not isinstance(filename, str):
            raise ValueError("Der Dateiname ('filename') muss ein String sein.")
        if not isinstance(table_name, str):
            raise ValueError("Der Tabellenname ('table_name') muss ein String sein.")
        if not isinstance(style_name, str):
            raise ValueError("Der Stilname ('style_name') muss ein String sein.")
        if not isinstance(freeze_first_row, bool):
            raise ValueError("Der Parameter ('freeze_first_row') muss ein Boolean sein.")
            
        if not os.path.isfile(filename):
            raise FileNotFoundError(f"Die Datei '{filename}' wurde nicht gefunden.")
            
        workbook = load_workbook(filename)
        sheet = workbook.active
        
        # Check if sheet has data
        if sheet.max_row == 1 and sheet.max_column == 1 and sheet['A1'].value is None:
            raise ValueError("Das Arbeitsblatt ist leer und kann nicht als Tabelle formatiert werden.")
        
        # Check if table name already exists
        existing_tables = [table.name for table in sheet._tables]
        if table_name in existing_tables:
            raise ValueError(f"Eine Tabelle mit dem Namen '{table_name}' existiert bereits.")
            
        table_ref = f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name=style_name, showFirstColumn=False, 
                                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
        
        if freeze_first_row:
            sheet.freeze_panes = "A2"
            
        workbook.save(filename)
        screen_and_log(f"Info: Datei '{filename}' erfolgreich als Tabelle formatiert.", logfile, screen)
        return True
        
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Eingabe: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Formatieren der Datei '{filename}': {e}", logfile, screen)
        return False

# -------------------------------
# Formatiert Zellen und Spaltenbreite einer existierenden EXCEL Datei
# -------------------------------


def format_excel_columns(filename, column_formats, column_widths=None, logfile=None, screen=True):
    """
    Öffnet eine Excel-Datei, formatiert die Spalten und passt deren Breite an.
    wenn nicht ausreichend formatiertungsangaben vorliegen bzw. übergeben werden, die letzte Spalteninformation für die folgenden Spalten verwendet  wird.

    Parameter:
        filename (str): Pfad zur Excel-Datei.
        column_formats (list): Liste von Formatstrings für Spalten (z. B. "DD.MM.YY", "#,##0.00").
        column_widths (list, optional): Liste von Breiten je Spalte. Wird die Liste überschritten,
                                        wird die letzte Breite wiederverwendet.
        logfile (str, optional): Optionaler Pfad zu einer Logdatei.
        screen (bool): Statusausgabe auf dem Bildschirm.
        
    Rückgabe:
        bool: True bei erfolgreichem Formatieren, False bei Fehler.
    """

    try:
        # Typprüfungen
        if not isinstance(filename, str):
            raise ValueError("Der Dateiname ('filename') muss ein String sein.")
        if not isinstance(column_formats, list) or not all(isinstance(fmt, str) for fmt in column_formats):
            raise ValueError("Die Spaltenformate ('column_formats') müssen eine Liste von Strings sein.")
        if column_widths and (not isinstance(column_widths, list) or not all(isinstance(w, (int, float)) for w in column_widths)):
            raise ValueError("Die Spaltenbreiten ('column_widths') müssen eine Liste von Zahlen sein.")
        if logfile and not isinstance(logfile, str):
            raise ValueError("Der Logfile-Pfad ('logfile') muss ein String sein.")

        if not os.path.isfile(filename):
            raise FileNotFoundError(f"Die Datei '{filename}' wurde nicht gefunden.")

        workbook = load_workbook(filename)
        sheet = workbook.active
        max_col = sheet.max_column

        for col_index in range(1, max_col + 1):
            # Ermittele gültige Format-Index
            fmt_index = min(col_index - 1, len(column_formats) - 1)
            fmt = column_formats[fmt_index]

            col_letter = sheet.cell(row=1, column=col_index).column_letter

            for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index, max_row=sheet.max_row):
                for cell in row:
                    cell.number_format = fmt

            if column_widths:
                width_index = min(col_index - 1, len(column_widths) - 1)
                sheet.column_dimensions[col_letter].width = column_widths[width_index]

        workbook.save(filename)
        screen_and_log(f"Info: Datei '{filename}' wurde erfolgreich formatiert und angepasst.", logfile, screen)
        return True

    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Datei '{filename}' wurde nicht gefunden: {e}", logfile, screen)
        return False
    except PermissionError:
        screen_and_log(f"ERROR: Keine Schreibberechtigung für die Datei '{filename}'.", logfile, screen)
        return False
    except ValueError as e:
        screen_and_log(f"ERROR: Ungültige Parameter: {e}", logfile, screen)
        return False
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Formatieren der Datei '{filename}': {e}", logfile, screen)
        return False



# --------------------------
# Parquet-Dateien importieren
# --------------------------
def import_parquet(filename, logfile=None, screen=True):
    """
    Liest eine Parquet-Datei ein und gibt einen DataFrame zurück.

    Parameter:
        filename (str): Pfad zur Parquet-Datei.
        logfile (str, optional): Der Pfad zum Logfile.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        pandas.DataFrame | None: Der DataFrame oder None bei Fehlern.
    """
    try:
        if not filename.endswith('.parquet'):
            raise ValueError(f"Die Datei '{filename}' hat keine '.parquet'-Endung.")
        if not os.path.isfile(filename):
            raise FileNotFoundError(f"Die Datei '{filename}' wurde nicht gefunden.")
        df = pd.read_parquet(filename)
        screen_and_log(f"Info: Parquet-Datei '{filename}' erfolgreich eingelesen.", logfile, screen)
        return df
    except Exception as e:
        screen_and_log(f"ERROR: Fehler beim Importieren der Datei '{filename}': {e}", logfile, screen)
        return None

# ----------------------------------------------
# Funktion: Prüfen, ob eine Datei gesperrt ist
# ----------------------------------------------
def is_file_open_windows(file_path):
    """
    Überprüft, ob eine Datei unter Windows gesperrt ist.

    Parameter:
        file_path (str): Pfad der Datei.

    Rückgabe:
        bool: True, wenn die Datei gesperrt ist, False sonst.
    """
    if os.name != 'nt':
        raise OSError("Diese Funktion unterstützt nur Windows.")
    if not isinstance(file_path, str):
        raise ValueError("file_path muss ein String sein.")
    if not os.path.exists(file_path):
        return False  # Datei nicht vorhanden

    try:
        with open(file_path, 'r+b') as file:
            try:
                msvcrt.locking(file.fileno(), msvcrt.LK_NBLCK, 1)
                msvcrt.locking(file.fileno(), msvcrt.LK_UNLCK, 1)
                return False
            except IOError:
                return True
    except Exception as e:
        raise RuntimeError(f"Ein unerwarteter Fehler ist aufgetreten: {e}")

# -------------------------------
# Zentralisierte Protokollierung
# -------------------------------
def screen_and_log(message, logfile=None, screen=True, auto_log=False):
    """
    Gibt die Nachricht auf dem Bildschirm aus (je nach Bedingungen) und schreibt sie optional in ein Logfile.

    Letzte Änderung: 03.01.25

    Parameter:
        message (str): Die Nachricht, die verarbeitet werden soll.
        logfile (str, optional): Der Pfad zum Logfile. Wenn None und auto_log=True, wird automatisch 'scriptname.log' verwendet.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden sollen.
        auto_log (bool): Wenn True und logfile=None, wird automatisch ein Logfile basierend auf dem Skriptnamen erstellt.

    Rückgabe:
        str: Die formatierte Nachricht (nützlich für Tests).
    """
    if not isinstance(message, str):
        raise ValueError("Die Nachricht ('message') muss ein String sein.")
    if logfile is not None and not isinstance(logfile, str):
        raise ValueError("Der Logfile-Pfad ('logfile') muss ein String sein, wenn angegeben.")
    
    # Auto-Logfile erstellen wenn gewünscht und kein Logfile angegeben
    if logfile is None and auto_log:
        # Ermittle den Namen des ursprünglichen Skripts (nicht dieser Bibliothek)
        stack = inspect.stack()
        script_filename = None
        for frame_info in reversed(stack):  # Von außen nach innen suchen
            filename = frame_info.filename
            if not filename.endswith('Standardfunktionen_aktuell.py'):
                script_filename = filename
                break
        
        if script_filename:
            script_name = os.path.splitext(os.path.basename(script_filename))[0]
            logfile = f"{script_name}.log"
    
    # Funktionsname des Aufrufers ermitteln
    caller_function = inspect.stack()[1].function
    
    # Zeitstempel ergänzen
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Nachricht formatieren
    formatted_message = f"{current_time} - {message} (Caller Funktion: {caller_function})"

    # Bildschirm-Ausgabe steuern
    if message.strip().upper().startswith(("ERROR", "WARNING")):
        # Fall 1: Immer formatierte Nachricht anzeigen (formatierte Ausgabe)
        print(formatted_message)
    elif screen is True:
        # Fall 2a: Originale Nachricht anzeigen, wenn `screen=True`
        print(message)
    # Fall 2b: Keine Anzeige, screen=False und keine "ERROR"/"WARNING"


    # Nachricht ins Logfile schreiben, falls ein Logfile angegeben ist
    if logfile:
        try:
            log_dir = os.path.dirname(logfile) or os.getcwd()
            os.makedirs(log_dir, exist_ok=True)  # Verzeichnis erstellen, falls es nicht existiert
            with open(logfile, 'a', encoding='utf-8') as log_file:
                log_file.write(formatted_message + "\n")
        except Exception as e:
            print(f"ERROR: Konnte nicht ins Logfile '{logfile}' schreiben: {e}")
    
    return formatted_message  # Optional: Rückgabe für Tests

# -------------------------------
# Arbeitsverzeichnis setzen
# -------------------------------
def set_working_directory(path="default", logfile=None, screen=True):
    """
    Setzt das aktuelle Arbeitsverzeichnis.

    Parameter:
        path (str): 
            - Wenn "default", wird das Arbeitsverzeichnis auf das Verzeichnis gesetzt, 
              von dem aus das Skript gestartet wurde.
            - Andernfalls wird der übergebene Pfad als Arbeitsverzeichnis verwendet.
        logfile (str, optional): Der Pfad zum Logfile. Wenn None, wird keine Protokollierung ins Logfile durchgeführt.
        screen (bool): Ob Nachrichten auf dem Bildschirm angezeigt werden.
        
    Rückgabe:
        None

    Hinweise:
        - Falls der Pfad ungültig ist, wird eine Fehlermeldung ausgegeben und das Verzeichnis nicht geändert.
        - Änderungen des Arbeitsverzeichnisses werden protokolliert.
    """
    try:
        # Typprüfung des Eingabeparameters
        if not isinstance(path, str):
            raise ValueError("Der Pfad ('path') muss ein String sein.")

        # Setzt das Arbeitsverzeichnis auf das Verzeichnis, aus dem das Caller Skript gestartet wurde
        if path == "default":
            # Ermitteln des Verzeichnisses des Callers
            caller_frame = inspect.stack()[1]
            caller_filename = caller_frame.filename
            caller_directory = os.path.dirname(os.path.abspath(caller_filename))
            os.chdir(caller_directory)
            screen_and_log(f"Info: Arbeitsverzeichnis wurde auf das Verzeichnis des Callers gesetzt: {caller_directory}", logfile, screen)
        else:
            # Prüft, ob der angegebene Pfad existiert und ein Verzeichnis ist
            if os.path.exists(path) and os.path.isdir(path):
                os.chdir(path)
                screen_and_log(f"Info: Arbeitsverzeichnis wurde auf '{path}' gesetzt.", logfile, screen)
            else:
                raise FileNotFoundError(f"Der Pfad '{path}' ist nicht verfügbar oder kein gültiges Verzeichnis.")

    except PermissionError:
        screen_and_log(f"ERROR: Keine Berechtigung, das Arbeitsverzeichnis auf '{path}' zu setzen.", logfile, screen)
    except FileNotFoundError as e:
        screen_and_log(f"ERROR: Der angegebene Pfad ist ungültig: {e}", logfile, screen)
    except Exception as e:
        screen_and_log(f"ERROR: Ein unerwarteter Fehler ist aufgetreten: {e}", logfile, screen)

# -------------------------------
# Einstellungen aus INI-Datei laden
# -------------------------------
import os
import configparser
import ast  # Ermöglicht sicheres Parsen von Python-Literalen (z. B. dict, list, int)

def settings_import(file_name):
    """
    Liest eine INI-Datei ein und gibt die Inhalte als Dictionary zurück.
    Unterstützt strukturierte Werte wie Dictionaries in einer Zeile (z. B. für Export-Optionen).

    Formatbeispiel in der INI:
        [Export]
        values_month_to_excel = {"enabled": true, "filename": "file.xlsx", "column_formats": ["DD.MM.YY"], "column_widths": [12]}

    Rückgabe:
        dict: Strukturierte Einstellungen, z. B. settings["Export"]["values_month_to_excel"]["filename"]
    """
    
    try:
        # Input validation
        if not isinstance(file_name, str):
            raise ValueError("Der Dateiname ('file_name') muss ein String sein.")
            
        # Prüfe, ob die Datei existiert
        if not os.path.isfile(file_name):
            raise FileNotFoundError(f"Die Datei '{file_name}' wurde nicht gefunden.")

        # Lade die INI-Datei
        config = configparser.ConfigParser(interpolation=None)
        config.read(file_name)

        # Zieldatenstruktur: Dictionary mit Abschnittsnamen (z. B. "Export") als Schlüssel
        settings = {}

        # Schleife über alle Abschnitte der INI-Datei
        for section in config.sections():
            settings[section] = {}  # Neuen Abschnitt vorbereiten

            # Schleife über alle Schlüssel/Wert-Paare in diesem Abschnitt
            for key, value in config.items(section):
                value = value.strip()  # Whitespace am Anfang/Ende entfernen
                
                # Versuche, strukturierte Werte (Dictionaries, Listen etc.) als Python-Objekt zu parsen
                if value.startswith("{") and value.endswith("}"):
                    try:
                        parsed = ast.literal_eval(value)  # Sicheres Parsen von Python-Datentypen
                        settings[section][key] = parsed   # Im Dictionary speichern
                        continue  # Weiter zum nächsten Eintrag
                    except Exception as e:
                        # Parsing fehlgeschlagen → als Fallback weitermachen
                        screen_and_log(f"WARNING: Kann Wert für '{section}:{key}' nicht als Dictionary parsen: {e}", 
                                     logfile, screen, auto_log=True)

                # Bool-Werte erkennen und umwandeln
                if value.lower() in ['true', 'false']:
                    settings[section][key] = value.lower() == 'true'
                # Ganzzahl oder Kommazahl erkennen
                elif value.replace('.', '', 1).isdigit():
                    settings[section][key] = float(value) if '.' in value else int(value)
                # Kommagetrennte Liste → z. B. "A,B,C"
                elif ',' in value:
                    settings[section][key] = [v.strip() for v in value.split(',')]
                else:
                    # Fallback: einfacher String
                    settings[section][key] = value

        return settings

    except Exception as e:
        # Allgemeine Fehlerbehandlung (z. B. Datei beschädigt)
        screen_and_log(f"ERROR: Fehler beim Laden der Einstellungen: {e}", logfile, screen, auto_log=True)
        return None
